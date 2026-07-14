import re
import cv2
import os
from pathlib import Path

import openpyxl

from qc_vision_pipeline_2 import transform_gt_bbox_homography_zahid2
from utils import load_detections_from_yolo, load_gt_from_labelme
from ultralytics import YOLO

os.environ["QT_QPA_FONTDIR"] = "/usr/share/fonts"
os.environ["QT_QPA_PLATFORM"] = "xcb"

# ======================================================
# CONFIGURATION -- edit these for each run
# ======================================================

# Point this at ONE camera folder at a time, e.g. ".../target/7"
# The folder name itself (must be a plain number, e.g. "7") is used both to
# find the matching GT image/json AND as the Excel column (1-42).
TARGET_FOLDER = "/home/mohammad/projects/smart_factory/sample_dataset/target/42"

GT_ROOT = "../sample_dataset//Ground_truth_v2"
OUTPUT_ROOT = "../qc_vision_results"

# The single Excel report that gets updated in place, run after run.
EXCEL_PATH = "../Inferr Results.xlsx"
EXCEL_SHEET_NAME = "Sheet1"
TS_COLUMN = 43          # column AQ
DATA_FIRST_ROW = 2      # row 1 is the header
DATA_LAST_ROW = 144     # 143 image rows

DISPLAY = False
SAVE_RESULTS = True

# Matches the "D20250205T171040" part out of "C0040_D20250205T171040.png"
TIMESTAMP_RE = re.compile(r"(D\d{8}T\d{6})")


def collect_images(folder: Path):
    """Collect all image files from a folder."""
    images = []
    images.extend(sorted(folder.glob("*.png")))
    images.extend(sorted(folder.glob("*.jpg")))
    images.extend(sorted(folder.glob("*.jpeg")))
    return images


def extract_timestamp(filename: str):
    """Pulls 'D20250205T171040' out of a filename like 'C0040_D20250205T171040.png'."""
    match = TIMESTAMP_RE.search(filename)
    return match.group(1) if match else None


def build_ts_row_map(ws, ts_col, row_start, row_end):
    """Maps each TS-column value (e.g. 'D20250205T171040') to its row number."""
    row_map = {}
    for row in range(row_start, row_end + 1):
        ts_value = ws.cell(row=row, column=ts_col).value
        if ts_value:
            row_map[str(ts_value).strip()] = row
    return row_map


def main():

    target_folder = Path(TARGET_FOLDER)
    folder_name = target_folder.name

    if not folder_name.isdigit():
        raise ValueError(
            f"TARGET_FOLDER's name must be a plain camera number "
            f"(e.g. '7'), got '{folder_name}'"
        )

    camera_index = int(folder_name)  # 1-42, also the Excel column number

    print("=" * 80)
    print(f"Processing camera folder: {folder_name} (Excel column {camera_index})")
    print("=" * 80)

    # ======================================================
    # Load YOLO
    # ======================================================

    print("Loading YOLO model...")
    model = YOLO("../models/SIDEOTRRH_MQ4_0.pt")

    # ======================================================
    # Find Corresponding GT Image + JSON
    # ======================================================

    gt_image_path = Path(GT_ROOT) / f"{folder_name}.png"
    if not gt_image_path.exists():
        gt_image_path = Path(GT_ROOT) / f"{folder_name}.jpg"
    if not gt_image_path.exists():
        gt_image_path = Path(GT_ROOT) / f"{folder_name}.jpeg"

    gt_json_path = Path(GT_ROOT) / f"{folder_name}.json"

    if not gt_image_path.exists():
        raise FileNotFoundError(f"GT image not found for folder {folder_name}")
    if not gt_json_path.exists():
        raise FileNotFoundError(f"GT json not found for folder {folder_name}")

    img_gt = cv2.imread(str(gt_image_path))
    if img_gt is None:
        raise FileNotFoundError(f"Failed to load GT image: {gt_image_path}")

    gt_boxes, hole_ids = load_gt_from_labelme(str(gt_json_path))
    print(f"GT Boxes: {len(gt_boxes)}")

    # ======================================================
    # Collect Target Images
    # ======================================================

    target_images = collect_images(target_folder)
    print(f"Target Images: {len(target_images)}")

    if len(target_images) == 0:
        print(f"No images found in {target_folder}")
        return

    # ======================================================
    # Output Folder For This Camera
    # ======================================================

    output_folder = Path(OUTPUT_ROOT) / folder_name
    output_folder.mkdir(parents=True, exist_ok=True)

    # ======================================================
    # Open The Existing Excel Report
    # ======================================================

    if not Path(EXCEL_PATH).exists():
        raise FileNotFoundError(
            f"Excel report not found at {EXCEL_PATH} -- this script updates "
            f"an existing file, it does not create a new one."
        )

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[EXCEL_SHEET_NAME]

    ts_row_map = build_ts_row_map(ws, TS_COLUMN, DATA_FIRST_ROW, DATA_LAST_ROW)

    rows_written = 0
    rows_skipped = 0

    # ======================================================
    # Process Each Image In The Folder
    # ======================================================

    for idx, target_path in enumerate(target_images):

        print(f"\n[{idx + 1}/{len(target_images)}] {target_path.name}")

        ts = extract_timestamp(target_path.name)
        if ts is None:
            print(f"  Could not extract a timestamp from '{target_path.name}', skipping.")
            rows_skipped += 1
            continue

        row = ts_row_map.get(ts)
        if row is None:
            print(f"  Timestamp '{ts}' not found in the TS column, skipping.")
            rows_skipped += 1
            continue

        img_target = cv2.imread(str(target_path))
        if img_target is None:
            print(f"  Failed to load: {target_path}")
            rows_skipped += 1
            continue

        try:
            detections = load_detections_from_yolo(model, str(target_path))
            print(f"  Detections: {len(detections)}")

            gt_bbox_dict = {0: gt_boxes}
            hole_id_dict = {0: hole_ids}
            detection_dict = {0: detections}

            output_filename = f"{target_path.stem}_icp{target_path.suffix}"

            # total_cams=1 here on purpose: this call handles a single GT/target
            # image pair (cam_id 0). It is NOT the physical camera number --
            # that's `camera_index`, used only for picking the Excel column.
            _, _, unmatched_hole_id_dict = transform_gt_bbox_homography_zahid2(
                total_cams=1,
                gt_bbox_dict=gt_bbox_dict,
                detection_dict=detection_dict,
                hole_id_dict=hole_id_dict,
                img_dict={0: {"gt": img_gt, "target": img_target}},
                DISPLAY=DISPLAY,
                SAVE_RES=SAVE_RESULTS,
                outdir=str(output_folder),
                output_filename=output_filename,
            )

            unmatched_ids = sorted(unmatched_hole_id_dict.get(0, []))
            cell_value = str(unmatched_ids)  # "[]" or "[37, 22]"

            ws.cell(row=row, column=camera_index, value=cell_value)
            rows_written += 1

            print(f"  Row {row} <- column {camera_index}: {cell_value}")

            if DISPLAY:
                cv2.waitKey(0)
                cv2.destroyAllWindows()

        except Exception as e:
            print(f"  Error processing {target_path.name}: {e}")
            rows_skipped += 1
            continue

        # Save after every image so a crash / early Ctrl-C never loses more
        # than the image currently being processed.
        wb.save(EXCEL_PATH)

    cv2.destroyAllWindows()

    print("\n" + "=" * 80)
    print(f"Folder {folder_name} done. Rows written: {rows_written}, skipped: {rows_skipped}")
    print(f"Excel report updated in place: {EXCEL_PATH}")
    print("=" * 80)


if __name__ == "__main__":
    main()
